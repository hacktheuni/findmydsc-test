from django.shortcuts import render, redirect
from user.models import *
from django.contrib import messages
import os, re
from django.conf import settings
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.urls import reverse
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.db.models import Count, Q
from django.contrib.auth.hashers import check_password, make_password
from django.utils import timezone
from datetime import timedelta
import openpyxl
from openpyxl.utils import get_column_letter
from user.views import allow_only_client_users, getUser

# User All Function are here for SubAdmin
@allow_only_client_users
def listUser(request):
    subAdminID = request.session.get('subAdminID')
    if subAdminID:
        try:
            subAdmin = SignUP.objects.get(subAdminID=subAdminID)
            user = UpdatedUser.objects.get(userPhone=subAdmin.subAdminPhone, isActive=False)
            users = UpdatedUser.objects.filter(subAdminID=subAdmin.subAdminID, isActive="True").all().order_by('-userModifiedDate')
            context = {
                'base': 'base/subAdminBase.html',
                'users': users,
                'user': user,
                'subAdmin': subAdmin
            }
            return render(request, 'user/listUser.html', context)
        except SignUP.DoesNotExist:
            messages.error(request, "SubAdmin not found.")
            return redirect('adminSignIn')
        except UpdatedUser.DoesNotExist:
            messages.error(request, "User not found.")
            return redirect('adminSignIn')
    else:
        messages.error(request, "Only Admin have the permission.")
        return redirect('adminSignIn')

@allow_only_client_users
def addUser(request):
    subAdminID = request.session.get('subAdminID')
    if subAdminID:
        try:
            subAdmin = SignUP.objects.get(subAdminID=subAdminID)
            user = UpdatedUser.objects.get(userPhone=subAdmin.subAdminPhone, isActive=False)
        except SignUP.DoesNotExist:
            messages.error(request, "SubAdmin not found.")
            return redirect('adminSignIn')
        except UpdatedUser.DoesNotExist:
            messages.error(request, "User not found.")
            return redirect('adminSignIn')

        groups = UpdatedGroup.objects.filter(subAdminID=user.subAdminID).all()
        context = {
            'base': 'base/subAdminBase.html',
            'user': user,
            'subAdmin': subAdmin,
            'groups': groups
        }

        if request.method == 'POST':
            userName = request.POST.get('userName')
            userPhone = request.POST.get('userPhone')
            userUsername = request.POST.get('userUsername')
            userPassword = request.POST.get('userPassword')
            groupName = request.POST.get('groupName', '').strip()
            perm = request.POST.get('perm', '')

            accessToPendingWork = 'accessToPendingWork' in request.POST
            accessToAnnual = 'accessToAnnual' in request.POST
            accessToTrademark = 'accessToTrademark' in request.POST

            readOnly = False
            readWrite = False
            isClientUser = False

            form_data = {
                'userName': userName,
                'userPhone': userPhone,
                'userUsername': userUsername,
                'groupName': groupName,
                'perm': perm
            }

            isGroupNameFilled = bool(groupName)
            isAnyCheckboxChecked = accessToPendingWork or accessToAnnual or accessToTrademark
            isPermissionSelected = perm in ['readOnly', 'readWrite']

            # Combined validation logic
            if isGroupNameFilled or isAnyCheckboxChecked or perm:
                if not isGroupNameFilled:
                    messages.error(request, "Group name is required.")
                    context['form_data'] = form_data
                    return render(request, 'user/addUser.html', context)

                if not isAnyCheckboxChecked:
                    messages.error(request, "At least one page access permission is required.")
                    context['form_data'] = form_data
                    return render(request, 'user/addUser.html', context)

                if not isPermissionSelected:
                    messages.error(request, "Please select either Read Only or Read & Write as permission.")
                    context['form_data'] = form_data
                    return render(request, 'user/addUser.html', context)

            # Fetch group object only if groupName is provided
            group_obj = None
            if groupName:
                try:
                    group_obj = UpdatedGroup.objects.get(subAdminID=user.subAdminID, groupName=groupName)
                    if perm == "readOnly":
                        readOnly = True
                        isClientUser = True
                    elif perm == "readWrite":
                        readWrite = True
                        isClientUser = True
                except UpdatedGroup.DoesNotExist:
                    messages.error(request, "Group not found.")
                    context['form_data'] = form_data
                    return render(request, 'user/addUser.html', context)

            # Required fields check
            if not userName or not userPhone or not userUsername or not userPassword:
                messages.error(request, "All fields are required.")
                context['form_data'] = form_data
                return render(request, 'user/addUser.html', context)

            # 1. Name validation
            if not re.match(r'^[A-Za-z\s]+$', userName):
                messages.error(request, "Name can only contain letters and spaces.")
                context['form_data'] = form_data
                return render(request, 'user/addUser.html', context)

            # 2. Phone validation
            if not re.match(r'^\d{10}$', userPhone):
                messages.error(request, "Phone number must be exactly 10 digits.")
                context['form_data'] = form_data
                return render(request, 'user/addUser.html', context)

            # 3. Password validation
            if len(userPassword) < 8 or not re.search(r'[A-Za-z]', userPassword) or not re.search(r'\d', userPassword) or not re.search(r'[@$!%*?&#]', userPassword):
                messages.error(request, "Password must be at least 8 characters long and contain letters, numbers, and special characters (@, $, !, %, *, ?, &, #).")
                context['form_data'] = form_data
                return render(request, 'user/addUser.html', context)

            # 4. Uniqueness checks
            if UpdatedUser.objects.filter(subAdminID=user.subAdminID, userPhone=userPhone).exists():
                messages.error(request, "Phone number already exists.")
                context['form_data'] = form_data
                return render(request, 'user/addUser.html', context)

            if UpdatedUser.objects.filter(subAdminID=user.subAdminID, userUsername=userUsername).exists():
                messages.error(request, "Username already exists.")
                context['form_data'] = form_data
                return render(request, 'user/addUser.html', context)

            # Save the user
            new_user = UpdatedUser(
                subAdminID=user.subAdminID,
                userName=userName,
                userPhone=userPhone,
                userUsername=userUsername,
                userPassword=make_password(userPassword),
                groupID=group_obj,
                isClientUser=isClientUser,
                canReadOnly=readOnly,
                canReadWrite=readWrite,
                accessToPendingWork=accessToPendingWork,
                accessToAnnual=accessToAnnual,
                accessToTrademark=accessToTrademark,
            )
            new_user.save()

            # Save to history
            userHistory = HistoryUser(
                subAdminID=user.subAdminID,
                userID=new_user,
                userName=userName,
                userPhone=userPhone,
                userUsername=userUsername,
                userPassword=new_user.userPassword,
                userModifiedDate=new_user.userModifiedDate,
                groupID=new_user.groupID,
                isClientUser=user.isClientUser,
                canReadOnly=user.canReadOnly,
                canReadWrite=user.canReadWrite,
                accessToPendingWork=user.accessToPendingWork,
                accessToAnnual=user.accessToAnnual,
                accessToTrademark=user.accessToTrademark
            )
            userHistory.save()

            messages.success(request, "User added successfully.")
            return redirect('listUser')

        return render(request, 'user/addUser.html', context)
    else:
        messages.error(request, "Only Admin has permission.")
        return redirect('adminSignIn')

@allow_only_client_users
def updateUser(request, userID):
    subAdminID = request.session.get('subAdminID')
    if subAdminID:
        try:
            subAdmin = SignUP.objects.get(subAdminID=subAdminID)
            user = UpdatedUser.objects.get(subAdminID=subAdmin.subAdminID, userID=userID)
        except SignUP.DoesNotExist:
            messages.error(request, "SubAdmin not found.")
            return redirect('adminSignIn')
        except UpdatedUser.DoesNotExist:
            messages.error(request, "User not found.")
            return redirect('adminSignIn')

        userHistory = HistoryUser.objects.filter(subAdminID=subAdmin.subAdminID, userID=userID).order_by('-userModifiedDate')
        groups = UpdatedGroup.objects.filter(subAdminID=user.subAdminID).all()

        context = {
            'base': 'base/subAdminBase.html',
            'user': user,
            'userHistory': userHistory,
            'subAdmin': subAdmin,
            'groups': groups
        }

        if request.method == 'POST':
            userName = request.POST.get('userName')
            userPhone = request.POST.get('userPhone')
            userUsername = request.POST.get('userUsername')
            userPassword = request.POST.get('userPassword')
            groupName = request.POST.get('groupName')
            perm = request.POST.get('perm')
            accessToPendingWork = 'accessToPendingWork' in request.POST
            accessToAnnual = 'accessToAnnual' in request.POST
            accessToTrademark = 'accessToTrademark' in request.POST

            readOnly = False
            readWrite = False
            isClientUser = False

            form_data = {
                'userName': userName,
                'userPhone': userPhone,
                'userUsername': userUsername,
                'groupName': groupName,
                'perm': perm
            }
            context['form_data'] = form_data

            group_obj = None
            if groupName:
                try:
                    group_obj = UpdatedGroup.objects.get(subAdminID=user.subAdminID, groupName=groupName)
                    if perm == "readOnly":
                        readOnly = True
                        isClientUser = True
                    elif perm == "readWrite":
                        readWrite = True
                        isClientUser = True
                except UpdatedGroup.DoesNotExist:
                    messages.error(request, "Group not found.")
                    return render(request, 'user/updateUser.html', context)

                if not accessToAnnual and not accessToPendingWork and not accessToTrademark:
                    messages.error(request, "At least one page access is required.")
                    return render(request, 'user/updateUser.html', context)

            # Empty field check
            if not userName or not userPhone or not userUsername:
                messages.error(request, "All fields are required.")
                return render(request, 'user/updateUser.html', context)

            # Name validation
            if not re.match(r'^[A-Za-z\s]+$', userName):
                messages.error(request, "Name can only contain letters and spaces.")
                return render(request, 'user/updateUser.html', context)

            # Phone validation
            if not re.match(r'^\d{10}$', userPhone):
                messages.error(request, "Phone number must be exactly 10 digits.")
                return render(request, 'user/updateUser.html', context)

            # Password validation (if updating)
            if userPassword:
                if len(userPassword) < 8 or not re.search(r'[A-Za-z]', userPassword) or not re.search(r'\d', userPassword) or not re.search(r'[@$!%*?&#]', userPassword):
                    messages.error(request, "Password must be at least 8 characters long and contain letters, numbers, and special characters (@, $, !, %, *, ?, &, #).")
                    return render(request, 'user/updateUser.html', context)

            # Uniqueness checks
            if UpdatedUser.objects.filter(subAdminID=subAdmin.subAdminID, userPhone=userPhone).exclude(userID=userID).exists():
                messages.error(request, "Phone number already exists.")
                return render(request, 'user/updateUser.html', context)

            if UpdatedUser.objects.filter(subAdminID=subAdmin.subAdminID, userUsername=userUsername).exclude(userID=userID).exists():
                messages.error(request, "Username already exists.")
                return render(request, 'user/updateUser.html', context)

            # Update user
            user.userName = userName
            user.userPhone = userPhone
            user.userUsername = userUsername
            if userPassword:
                user.userPassword = make_password(userPassword)
            user.groupID = group_obj
            user.isClientUser = isClientUser
            user.canReadOnly = readOnly
            user.canReadWrite = readWrite
            user.accessToPendingWork = accessToPendingWork
            user.accessToAnnual = accessToAnnual
            user.accessToTrademark = accessToTrademark
            user.save()

            # Add history
            userHistory = HistoryUser(
                subAdminID=user.subAdminID,
                userID=user,
                userName=user.userName,
                userPhone=user.userPhone,
                userUsername=user.userUsername,
                userPassword=user.userPassword,
                userModifiedDate=user.userModifiedDate,
                groupID=user.groupID,
                isClientUser=user.isClientUser,
                canReadOnly=user.canReadOnly,
                canReadWrite=user.canReadWrite,
                accessToPendingWork=user.accessToPendingWork,
                accessToAnnual=user.accessToAnnual,
                accessToTrademark=user.accessToTrademark
            )
            userHistory.save()

            messages.success(request, "User updated successfully.")
            return redirect(request.path)

        return render(request, 'user/updateUser.html', context)

    else:
        messages.error(request, "Only Admin has the permission.")
        return redirect('adminSignIn')

@allow_only_client_users
def deleteUser(request):
    if request.method == 'POST':
        userIDs = request.POST.getlist('userIDs')
        confirmation = request.POST.get('deleteUser')
        if confirmation:
            if userIDs:
                # Retrieve users to deactivate that are still active
                users_to_deactivate = UpdatedUser.objects.filter(userID__in=userIDs, isActive=True)
                
                if users_to_deactivate.exists():
                    # Deactivating users and setting the deactivatedBy field to 'subAdmin'
                    users_to_deactivate.update(isActive=False, deactivatedBy='subAdmin')
                    messages.success(request, "Selected users have been deactivated successfully.")
                else:
                    messages.error(request, "No active users were found to deactivate.")
            else:
                messages.error(request, "No users selected for deactivation.")
        else:
            messages.error(request, "Deletion not confirmed.")

    return redirect('listUser')


# All profile related Function are here for subAdmin

@allow_only_client_users 
def updateProfile(request):
    user_data = getUser(request)
    user = user_data.get('user')
    subAdmin = user_data.get('subAdmin')
    base = user_data.get('base')
    if subAdmin:
        context = {
            'base': base,
            'subAdmin': subAdmin,
            'user': user,
            'options': ["Company/ LLP", "Chartered Accountant", "Company Secretary", "Cost Accountant", "Others"]
        }

        if request.method == 'POST':
            # Retrieve form data
            subAdminName = request.POST.get('subAdminName')
            subAdminType = request.POST.get('subAdminType')
            subAdminEmail = request.POST.get('subAdminEmail')
            subAdminPhone = request.POST.get('subAdminPhone')
            subAdminCity = request.POST.get('subAdminCity')
            subAdminState = request.POST.get('subAdminState')
            subAdminPinCode = request.POST.get('subAdminPinCode')

            # Handle file upload if a new logo is provided
            if 'subAdminLogo' in request.FILES:
                logo = request.FILES['subAdminLogo']

                # Validate file size (max 500KB)
                if logo.size > 500 * 1024:  # 500KB
                    messages.error(request, "The logo file is too large. Maximum size allowed is 500KB.")
                    return render(request, 'adminDetails/updateProfile.html', context)

                # Validate file type (only PNG and JPEG)
                file_ext = os.path.splitext(logo.name)[1].lower()
                if file_ext not in ['.png', '.jpg', '.jpeg']:
                    messages.error(request, "Invalid file format. Only PNG and JPEG files are allowed.")
                    return render(request, 'adminDetails/updateProfile.html', context)

                # Check if an old logo exists and delete it
                if subAdmin.subAdminLogo:
                    old_logo_path = os.path.join(settings.MEDIA_ROOT, subAdmin.subAdminLogo.name)
                    if os.path.exists(old_logo_path):
                        os.remove(old_logo_path)

                # Save the new logo
                subAdmin.subAdminLogo = logo

            # Update the rest of the subAdmin fields
            subAdmin.subAdminName = subAdminName
            subAdmin.subAdminType = subAdminType
            subAdmin.subAdminEmail = subAdminEmail
            subAdmin.subAdminPhone = subAdminPhone
            subAdmin.subAdminCity = subAdminCity
            subAdmin.subAdminState = subAdminState
            subAdmin.subAdminPinCode = subAdminPinCode
            subAdmin.save()
            messages.success(request, 'Profile Updated Successfully.')
            return redirect('updateProfile')  # Redirect after successful update

        return render(request, 'adminDetails/updateProfile.html', context)
    else:
        messages.error(request, "Only Admin have the permission.")
        return redirect('adminSignIn')

@allow_only_client_users
def deleteProfile(request):
    subAdmin = getUser(request).get('subAdmin')
    if request.method == 'POST':
        deleteProfile = request.POST.get('deleteProfile')
        subAdminPassword = request.POST.get('subAdminPassword')

        if deleteProfile:
            try:
                # Check if the provided password matches the hashed password
                if check_password(subAdminPassword, subAdmin.subAdminPassword):
                    subAdmin.delete()  # Delete the profile if the password matches
                    # Clear all session data
                    request.session.flush()
                    messages.success(request, "Your Account is Deleted.")
                    return redirect('adminSignIn')
                else:
                    messages.error(request, "Password does not match.")
                    return redirect('updateProfile')

            except SignUP.DoesNotExist:
                messages.error(request, "Account not found.")
                return redirect('updateProfile')
        else:
            messages.error(request, "Deletion not confirmed.")
            return redirect('updateProfile')

    return redirect('updateProfile')

@allow_only_client_users
def subscriptionDetails(request):
    subAdminID = request.session.get('subAdminID')
    context = {
        'base': 'base/subAdminBase.html',
        'subAdmin': None,
        'user': None,
        'subscriptionPlan': None,
        'activePlan': None,
        'start_date': None,
        'end_date': None,
        'time_remaining': None,
        'formatted_time_remaining': None,  # For formatted output
    }

    if subAdminID:
        subAdmin = SignUP.objects.get(subAdminID=subAdminID)
        user = UpdatedUser.objects.get(userPhone=subAdmin.subAdminPhone, isActive=False)
        subscriptionPlan = SubscriptionPlan.objects.all()
        activePlan = SubAdminSubscription.objects.filter(subAdminID=subAdmin, isActive=True).first()

        if subAdmin.hasUsedFreePlan:
            subscriptionPlan = list(SubscriptionPlan.objects.all())
            subscriptionPlan.pop(0)

        # Add logic to get subscription details if an active plan exists
        if activePlan:
            context['activePlan'] = activePlan
            context['start_date'] = activePlan.startDate
            context['end_date'] = activePlan.endDate

            # Calculate time remaining
            time_remaining = activePlan.endDate - timezone.now()
            if time_remaining > timedelta(0):
                context['time_remaining'] = time_remaining
                # Format time remaining for display
                days, seconds = time_remaining.days, time_remaining.seconds
                hours, remainder = divmod(seconds, 3600)
                context['formatted_time_remaining'] = f"{days} days, {hours} hours"
            else:
                context['time_remaining'] = timedelta(0)  # Plan has expired
                context['formatted_time_remaining'] = "Subscription has expired"

        context['subAdmin'] = subAdmin
        context['user'] = user
        context['subscriptionPlan'] = subscriptionPlan
    else:
        messages.error(request, "Only Admin have the permission.")
        return redirect('adminSignIn')

    return render(request, 'adminDetails/subscriptionDetails.html', context)

@allow_only_client_users
def exportToExcel(request):
    sub_admin_id = request.session.get('subAdminID')
    
    if not sub_admin_id:
        messages.error(request, "Only Admins have permission.")
        return redirect('adminSignIn')

    subAdmin = SignUP.objects.get(subAdminID=sub_admin_id)
    filename = f"{subAdmin.subAdminName} data.xlsx"

    # Create an Excel workbook
    wb = openpyxl.Workbook()
    
    # ================================
    # 1 **DSC SHEET**
    # ================================
    ws1 = wb.active
    ws1.title = "DSCs"
    
    # Define headers for DSC data
    headers = [
        'Group Name', 'Company Name', 'Client Name', 'Status', 'Location', 
        'Renewal Date', 'Contact Person', 'Phone Number', 'Last Modified Date', 
        'Last Received By', 'Last Received From', 'Last Delivered By', 'Last Delivery To'
    ]
    ws1.append(headers)

    dsc_data = UpdatedDSC.objects.filter(subAdminID=sub_admin_id).select_related('companyID__groupID', 'userID')

    for dsc in dsc_data:
        company = dsc.companyID
        group = company.groupID if company else None
        client = UpdatedClient.objects.filter(companyID=company).first() if company else None

        ws1.append([
            group.groupName if group else '',
            company.companyName if company else '',
            dsc.clientName,
            dsc.status,
            dsc.location,
            dsc.renewalDate.strftime('%d-%m-%Y') if dsc.renewalDate else '',
            client.clientName if client else '',
            client.clientPhone if client else '',
            dsc.modifiedDate.strftime('%d-%m-%Y %H:%M:%S'),
            dsc.receivedBy,
            dsc.receivedFrom,
            dsc.deliveredBy,
            dsc.deliveredTo
        ])

    # ================================
    # 2 **Pending Work SHEET**
    # ================================
    ws2 = wb.create_sheet(title="Pending Work")
    ws2.append(['Company Name', 'Group Name', 'Form No.', 'Event Date', 'Status', 'Internal Due Date', 'Due Date', 'Remarks', 'SRN No.','SRN Date','SRN Amount', 'Responsible Person', 'Fees', 'Billing'])

    pending_works = PendingWork.objects.filter(subAdminID=sub_admin_id, isArchived=False).all()
    for task in pending_works:
        ws2.append([
            task.companyID.companyName,
            task.companyID.groupID.groupName,
            task.formID.formNo,
            task.eventDate,
            task.status,
            task.internalDueDate,
            task.actualDueDate,
            task.remark,
            task.srnNo,
            task.srnDate,
            task.amt,
            task.userID.userName,
            task.fees,
            task.billing,
        ])

    # ================================
    # 3 **Pending Work Archived SHEET**
    # ================================
    ws3 = wb.create_sheet(title="Pending Work Archived")
    ws3.append(['Company Name', 'Group Name', 'Form No.', 'Event Date', 'Status', 'Internal Due Date', 'Due Date', 'Remarks', 'SRN No.','SRN Date','SRN Amount', 'Responsible Person', 'Fees', 'Billing'])

    pending_works = PendingWork.objects.filter(subAdminID=sub_admin_id, isArchived=True).all()
    for task in pending_works:
        ws3.append([
            task.companyID.companyName,
            task.companyID.groupID.groupName,
            task.formID.formNo,
            task.eventDate,
            task.status,
            task.internalDueDate,
            task.actualDueDate,
            task.remark,
            task.srnNo,
            task.srnDate,
            task.amt,
            task.userID.userName,
            task.fees,
            task.billing,
        ])

    # ================================
    # 4 **Annual Filing SHEET**
    # ================================
    ws4 = wb.create_sheet(title="Annual Filing")
    ws4.append(['Company Name', 'Group Name', 'DPT-3', 'MGT-14', 'AOC-4', 'MGT-7', 'Form-11', 'Form-8', 'Financial Year'])

    annual_filings = AnnualFiling.objects.filter(subAdminID=sub_admin_id, isArchived=False).all()
    for filing in annual_filings:
        ws4.append([
            filing.companyID.companyName,
            filing.companyID.groupID.groupName,
            filing.statusDPT3,
            filing.statusMGT14,
            filing.statusAOC4,
            filing.statusMGT7,
            filing.statusForm11,
            filing.statusForm8,
            filing.financialYear,
        ])

    # ================================
    # 5 **Annual Filing Archived SHEET**
    # ================================
    ws5 = wb.create_sheet(title="Annual Filing Archived")
    ws5.append(['Company Name', 'Group Name', 'DPT-3', 'MGT-14', 'AOC-4', 'MGT-7', 'Form-11', 'Form-8', 'Financial Year'])

    annual_filings = AnnualFiling.objects.filter(subAdminID=sub_admin_id, isArchived=True).all()
    for filing in annual_filings:
        ws5.append([
            filing.companyID.companyName,
            filing.companyID.groupID.groupName,
            filing.statusDPT3,
            filing.statusMGT14,
            filing.statusAOC4,
            filing.statusMGT7,
            filing.statusForm11,
            filing.statusForm8,
            filing.financialYear,
        ])

    # ================================
    # 6 **Pending Work Report SHEET**
    # ================================
    ws6 = wb.create_sheet(title="Pending Work Report")
    ws6.append(['SRN Date', 'SRN No.', 'Company Name', 'Group Name', 'Status', 'Form No.', 'SRN Challan Amt'])

    pwReports = PendingWork.objects.filter(subAdminID=sub_admin_id).all()
    for report in pwReports:
        ws6.append([
            report.srnDate,
            report.srnNo,
            report.companyID.companyName,
            report.companyID.groupID.groupName,
            report.status,
            report.formID.formNo,
            report.amt,
        ])

    # ================================
    # 7 **Annual Filing Report SHEET**
    # ================================
    ws7 = wb.create_sheet(title="Annual Filing Report")
    ws7.append(['SRN Date', 'SRN No.', 'Company Name', 'Group Name', 'Status', 'Form No.', 'SRN Challan Amt'])

    afReports = AnnualFiling.objects.filter(subAdminID=sub_admin_id).all()

    for report in afReports:
        approved_forms = []  # List to store approved form details
        
        if report.statusDPT3 == "Approved":
            approved_forms.append((report.srnDateDPT3, report.srnNoDPT3, "DPT3", report.amtDPT3))

        if report.statusMGT14 == "Approved":
            approved_forms.append((report.srnDateMGT14, report.srnNoMGT14, "MGT14", report.amtMGT14))

        if report.statusAOC4 == "Approved":
            approved_forms.append((report.srnDateAOC4, report.srnNoAOC4, "AOC4", report.amtAOC4))

        if report.statusMGT7 == "Approved":
            approved_forms.append((report.srnDateMGT7, report.srnNoMGT7, "MGT7", report.amtMGT7))

        if report.statusForm11 == "Approved":
            approved_forms.append((report.srnDateForm11, report.srnNoForm11, "Form 11", report.amtForm11))

        if report.statusForm8 == "Approved":
            approved_forms.append((report.srnDateForm8, report.srnNoForm8, "Form 8", report.amtForm8))

        # Append only approved forms to the sheet
        for srn_date, srn_no, form_no, amt in approved_forms:
            ws7.append([
                srn_date,
                srn_no,
                report.companyID.companyName,
                report.companyID.groupID.groupName,
                "Approved",
                form_no,
                amt,
            ])

    # ================================
    # 8 **Trademark SHEET**
    # ================================
    ws8 = wb.create_sheet(title="Trademark")
    ws8.append(['SRN No.', 'Name of Trademark', 'Name of Applicant', 'Group', 'Application No.', 'Class ', 'Date of Application', 'Current Status 1', 'Current Status 2', 'Remarks', 'Notice Receive / Serve Date', 'Reply Due Date', 'Fees Amt.', 'Hearing Date', 'Renewal Date', 'Fees Received'])

    trademarks = Trademark.objects.filter(subAdminID=sub_admin_id, isArchived=False).all()
    for tm in trademarks:
        ws8.append([
            tm.indexSRN,
            tm.nameOfTrademark,
            tm.nameOfApplicant,
            tm.groupID.groupName,
            tm.applicationNo,
            tm.classNo,
            tm.dateOfApp,
            tm.status1,
            tm.status2,
            tm.remark,
            tm.oppDate,
            tm.lastDate,
            tm.fees,
            tm.hearingDate,
            tm.expiryDate,
            tm.feesStatus,
        ])



    # ================================
    # 9 **Trademark Archived SHEET**
    # ================================
    ws9 = wb.create_sheet(title="Trademark Archived")
    ws9.append(['SRN No.', 'Name of Trademark', 'Name of Applicant', 'Group', 'Application No.', 'Class ', 'Date of Application', 'Current Status 1', 'Current Status 2', 'Remarks', 'Notice Receive / Serve Date', 'Reply Due Date', 'Fees Amt.', 'Hearing Date', 'Renewal Date', 'Fees Received'])

    trademarks = Trademark.objects.filter(subAdminID=sub_admin_id, isArchived=True).all()
    for tm in trademarks:
        ws9.append([
            tm.indexSRN,
            tm.nameOfTrademark,
            tm.nameOfApplicant,
            tm.groupID.groupName,
            tm.applicationNo,
            tm.classNo,
            tm.dateOfApp,
            tm.status1,
            tm.status2,
            tm.remark,
            tm.oppDate,
            tm.lastDate,
            tm.fees,
            tm.hearingDate,
            tm.expiryDate,
            tm.feesStatus,
        ])

    

    # Auto-adjust column width for better readability
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            sheet.column_dimensions[col_letter].width = max_length + 2

    # ================================
    # ✅ **Prepare and return response**
    # ================================
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)

    return response

@allow_only_client_users            
def exportData(request):
    if request.session.get('subAdminID'):
        
        context = {
            'base': 'base/subAdminBase.html'
        }
    else:
        messages.error(request, "Only Admin have the permission.")
        return redirect('adminSignIn')
    return render(request, 'adminDetails/exportData.html', context)


# All Function are here for the superAdmin
def listSubAdmin(request):
    if request.session.get('superAdminID'):
        subAdmins = SignUP.objects.annotate(
            active_user_count=Count('updateduser', filter=Q(updateduser__isActive=True), distinct=True),
            dsc_count=Count('updateddsc', distinct=True)  # Ensure distinct DSC entries are counted
        ).order_by('-subAdminRegisterDate')
        context = {
            'subAdmins': subAdmins
        }
        return render(request, 'subAdmin/listSubAdmin.html', context)
    else:
        messages.error(request, "Only Admin have the permission.")
        return redirect('adminSignIn')

def listFeedback(request):
    if request.session.get('superAdminID'):
        feedbacks = Feedback.objects.all().order_by('-feedbackDate')
        context = {
            'feedbacks': feedbacks
        }
        return render(request, 'contactUs/listFeedback.html', context)
    else:
        messages.error(request, "Only Admin have the permission.")
        return redirect('adminSignIn')

def action(request):
    if request.method == 'POST':
        subAdminIDs = request.POST.getlist('subAdminIDs')
        action_type = request.POST.get('action_type')  

        if subAdminIDs:
            if action_type == 'deactivate':
                # Deactivate subAdmins and their users
                subAdmins_to_deactivate = SignUP.objects.filter(subAdminID__in=subAdminIDs, isActive=True)
                if subAdmins_to_deactivate.exists():
                    subAdmins_to_deactivate.update(isActive=False)  # Deactivating subAdmins
                    
                    # Deactivate users of the subAdmins
                    users_to_deactivate = UpdatedUser.objects.filter(subAdminID__in=subAdminIDs, isActive=True)
                    users_to_deactivate.update(isActive=False, deactivatedBy='superAdmin')  # Deactivating users and marking who deactivated them

                    messages.success(request, "Selected subAdmins and their users have been deactivated successfully.")
                else:
                    messages.error(request, "Some subAdmins are already deactivated or do not exist.")

            elif action_type == 'activate':
                # Activate subAdmins
                subAdmins_to_activate = SignUP.objects.filter(subAdminID__in=subAdminIDs, isActive=False)
                if subAdmins_to_activate.exists():
                    subAdmins_to_activate.update(isActive=True)  # Activating subAdmins

                    # Activate users of the subAdmins, excluding those deactivated by subAdmin
                    users_to_activate = UpdatedUser.objects.filter(subAdminID__in=subAdminIDs, isActive=False, deactivatedBy='superAdmin')
                    users_to_activate.update(isActive=True, deactivatedBy=None)  # Activating users and clearing the deactivation marker

                    messages.success(request, "Selected subAdmins and their eligible users have been activated successfully.")
                else:
                    messages.error(request, "Some subAdmins are already active or do not exist.")
        else:
            messages.error(request, "No subAdmins selected for activation or deactivation.")
        
        return redirect('listSubAdmin')

    return redirect('listSubAdmin')


